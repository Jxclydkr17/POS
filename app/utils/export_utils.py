# app/utils/export_utils.py
"""
FASE 3 — Fix 3.2: Exportaciones Excel con openpyxl puro.

Se eliminó la dependencia de pandas/numpy (~100MB en el .exe empaquetado).
openpyxl ya estaba en requirements.txt como dependencia de pandas; ahora
es la única librería necesaria para generar Excel.

Las funciones de exportación a PDF (reportlab) no cambiaron.
"""
from openpyxl import Workbook
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas
from datetime import datetime
from pathlib import Path
from app.utils.dt import now_cr
from app.core.config import DATA_DIR  # FASE 2 — Fix 2.3: exports en DATA_DIR


# ── FASE 2 — Fix 2.3: directorio de exports persistente ─────
# Antes: los exports se escribían con rutas relativas al CWD
# ("exports/foo.xlsx"). En una app empaquetada como .exe, el CWD puede
# ser cualquier sitio según cómo se lance (acceso directo del escritorio,
# menú inicio, doble-click directo, etc.) → el usuario nunca encontraba
# sus archivos. DATA_DIR/exports/ es predecible y persiste entre updates.

def _export_dir() -> Path:
    """Retorna DATA_DIR/exports/, creándolo si no existe."""
    p = DATA_DIR / "exports"
    p.mkdir(parents=True, exist_ok=True)
    return p


def _resolve_export_filename(filename, default_basename: str) -> str:
    """Resuelve cualquier filename a una ruta absoluta bajo DATA_DIR/exports/.

    Reglas:
      - None / vacío  → DATA_DIR/exports/<default_basename>
      - relativo      → DATA_DIR/exports/<basename(filename)>
                        (ignora prefijos como 'exports/' para evitar duplicación)
      - absoluto      → respeta tal cual; sólo asegura que el directorio padre
                        exista. El caller asumió la responsabilidad de elegir
                        el destino.
    Siempre retorna `str` (compatible con openpyxl/reportlab que esperan str).
    """
    if filename:
        p = Path(filename)
        if p.is_absolute():
            p.parent.mkdir(parents=True, exist_ok=True)
            return str(p)
        # Relativo: tomar sólo el basename para no duplicar 'exports/'
        return str(_export_dir() / p.name)
    return str(_export_dir() / default_basename)


# ── Helpers ─────────────────────────────────────────────────

def _write_sheet(ws, rows: list[dict], *, headers: list[str] | None = None):
    """Escribe una lista de dicts en una hoja openpyxl con headers automáticos."""
    if not rows:
        return
    cols = headers or list(rows[0].keys())
    ws.append(cols)
    for row in rows:
        ws.append([row.get(c, "") for c in cols])


def _dicts_to_excel(data: list[dict], filename: str) -> str:
    """Exporta una lista de dicts a un archivo .xlsx de una sola hoja."""
    wb = Workbook()
    ws = wb.active
    _write_sheet(ws, data)
    wb.save(filename)
    return filename


# ============================================================
# VENTAS — Exportaciones
# ============================================================

def export_sales_history_excel(data, filename=None):
    """Exporta lista de ventas a Excel"""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"historial_ventas_{now_cr().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return _dicts_to_excel(data, filename)


def export_sales_history_pdf(data, start_date, end_date, filename=None, business_name="Mi Negocio"):
    """Exporta lista de ventas a PDF"""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"historial_ventas_{now_cr().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    c = canvas.Canvas(filename, pagesize=landscape(letter))
    width, height = landscape(letter)
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(260, y, f"Histórico de Ventas - {business_name}")
    y -= 30
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Rango: {start_date} al {end_date}")
    y -= 15
    c.drawString(50, y, f"Generado: {now_cr().strftime('%d/%m/%Y %H:%M')}")
    y -= 30

    headers = ["#Venta", "Fecha", "Cliente", "Pago", "Total (₡)"]
    col_widths = [70, 120, 220, 100, 100]

    c.setFont("Helvetica-Bold", 10)
    x = 50
    for i, h in enumerate(headers):
        c.drawString(x, y, h)
        x += col_widths[i]
    y -= 15
    c.line(50, y + 5, width - 50, y + 5)

    c.setFont("Helvetica", 9)
    for row in data:
        if y < 50:
            c.showPage()
            y = height - 50
        x = 50
        vals = [row["id"], row["created_at"], row["customer_name"], row["payment_method"], f"₡{row['total']:,.2f}"]
        for i, v in enumerate(vals):
            c.drawString(x, y, str(v))
            x += col_widths[i]
        y -= 15

    c.showPage()
    c.save()
    return filename


# ============================================================
# GASTOS — Exportaciones
# ============================================================

def export_expenses_excel(data, filename=None):
    """Exporta lista de gastos a Excel"""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"reporte_gastos_{now_cr().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return _dicts_to_excel(data, filename)


def export_expenses_pdf(data, start_date, end_date, total, filename=None, business_name="Mi Negocio"):
    """Exporta lista de gastos a PDF"""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"reporte_gastos_{now_cr().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    c = canvas.Canvas(filename, pagesize=landscape(letter))
    width, height = landscape(letter)
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(250, y, f"Reporte de Gastos - {business_name}")
    y -= 30
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Rango: {start_date} al {end_date}")
    y -= 15
    c.drawString(50, y, f"Generado: {now_cr().strftime('%d/%m/%Y %H:%M')}")
    y -= 30

    headers = ["Fecha", "Categoría", "Descripción", "Monto (₡)", "Método"]
    col_widths = [90, 120, 250, 100, 100]

    c.setFont("Helvetica-Bold", 10)
    x = 50
    for i, h in enumerate(headers):
        c.drawString(x, y, h)
        x += col_widths[i]
    y -= 15
    c.line(50, y + 5, width - 50, y + 5)

    c.setFont("Helvetica", 9)
    for row in data:
        if y < 50:
            c.showPage()
            y = height - 50
        x = 50
        vals = [
            row["date"], row["category"], row["description"] or "—",
            f"₡{row['amount']:,.2f}", row["payment_method"] or "—"
        ]
        for i, v in enumerate(vals):
            c.drawString(x, y, str(v))
            x += col_widths[i]
        y -= 15

    y -= 20
    c.setFont("Helvetica-Bold", 11)
    c.drawString(400, y, f"Total de gastos: ₡{total:,.2f}")

    c.showPage()
    c.save()
    return filename


# ============================================================
# COMPRAS — Exportaciones
# ============================================================

def export_purchases_excel(data, filename=None):
    """Exporta lista de compras/facturas a Excel."""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"reporte_compras_{now_cr().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return _dicts_to_excel(data, filename)


def export_purchases_pdf(data, title_extra="", filename=None, business_name="Mi Negocio"):
    """Exporta lista de compras/facturas a PDF."""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"reporte_compras_{now_cr().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    c = canvas.Canvas(filename, pagesize=landscape(letter))
    width, height = landscape(letter)
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, y, f"Reporte de Compras - {business_name}")
    y -= 25
    if title_extra:
        c.setFont("Helvetica", 10)
        c.drawString(50, y, title_extra)
        y -= 15
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Generado: {now_cr().strftime('%d/%m/%Y %H:%M')}")
    y -= 30

    headers = ["#", "Factura", "Proveedor", "F.Entrada", "F.Venc.", "Monto", "Abonado", "Saldo", "Estado"]
    col_widths = [40, 80, 140, 80, 80, 80, 80, 80, 70]

    c.setFont("Helvetica-Bold", 9)
    x = 40
    for i, h in enumerate(headers):
        c.drawString(x, y, h)
        x += col_widths[i]
    y -= 12
    c.line(40, y + 5, width - 40, y + 5)

    c.setFont("Helvetica", 8)
    total_amount = 0.0
    total_balance = 0.0

    for row in data:
        if y < 50:
            c.showPage()
            y = height - 50

        amount = float(row.get("amount", 0))
        paid = float(row.get("paid_amount", 0))
        balance = float(row.get("balance", amount))
        total_amount += amount
        total_balance += balance

        x = 40
        vals = [
            row.get("id", ""),
            row.get("invoice_number", ""),
            str(row.get("supplier_name", row.get("supplier_id", "")))[:22],
            str(row.get("entry_date", ""))[:10],
            str(row.get("due_date", ""))[:10],
            f"₡{amount:,.2f}",
            f"₡{paid:,.2f}",
            f"₡{balance:,.2f}",
            row.get("status", ""),
        ]
        for i, v in enumerate(vals):
            c.drawString(x, y, str(v))
            x += col_widths[i]
        y -= 12

    y -= 20
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y, f"Total facturado: ₡{total_amount:,.2f}    |    Saldo pendiente: ₡{total_balance:,.2f}")

    c.showPage()
    c.save()
    return filename


# ============================================================
# ANALÍTICA DE VENTAS — Exportaciones
# ============================================================

def export_sales_analytics_excel(data, filename=None):
    """Exporta analítica de ventas a Excel (múltiples hojas)."""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"analitica_ventas_{now_cr().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb = Workbook()

    # Hoja 1: KPIs
    ws = wb.active
    ws.title = "KPIs"
    kpis = data.get("kpis") or {}
    compare = data.get("compare") or {}
    previous = compare.get("previous", {})
    kpi_rows = [
        {"Indicador": "Ventas totales", "Actual": kpis.get("total_sales", 0),
         "Periodo anterior": previous.get("count", "")},
        {"Indicador": "Monto total (₡)", "Actual": kpis.get("total_amount", 0),
         "Periodo anterior": previous.get("total_amount", "")},
        {"Indicador": "Ticket promedio (₡)", "Actual": kpis.get("avg_ticket", 0),
         "Periodo anterior": previous.get("avg_ticket", "")},
    ]
    _write_sheet(ws, kpi_rows)

    daily = data.get("daily") or []
    if daily:
        _write_sheet(wb.create_sheet("Ventas diarias"), daily)

    cats = data.get("categories") or []
    if cats:
        _write_sheet(wb.create_sheet("Por categoría"), cats)

    top = data.get("top_products") or []
    if top:
        _write_sheet(wb.create_sheet("Top productos"), top)

    payments = data.get("payments") or []
    if payments:
        _write_sheet(wb.create_sheet("Métodos de pago"), payments)

    wb.save(filename)
    return filename


def export_sales_analytics_pdf(data, filename=None, business_name="Mi Negocio"):
    """Exporta analítica de ventas a PDF."""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"analitica_ventas_{now_cr().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    c = canvas.Canvas(filename, pagesize=landscape(letter))
    width, height = landscape(letter)
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, y, f"Analítica de Ventas - {business_name}")
    y -= 25
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Periodo: {data.get('start_date', '')} al {data.get('end_date', '')}")
    y -= 15
    c.drawString(50, y, f"Generado: {now_cr().strftime('%d/%m/%Y %H:%M')}")
    y -= 30

    kpis = data.get("kpis") or {}
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Indicadores Clave")
    y -= 20
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Ventas totales: {kpis.get('total_sales', 0)}")
    y -= 15
    c.drawString(50, y, f"Monto total: ₡{kpis.get('total_amount', 0):,.2f}")
    y -= 15
    c.drawString(50, y, f"Ticket promedio: ₡{kpis.get('avg_ticket', 0):,.2f}")
    y -= 30

    top = data.get("top_products") or []
    if top:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Top Productos")
        y -= 20
        headers = ["#", "Producto", "Cantidad", "Total (₡)"]
        col_widths = [30, 300, 80, 100]
        c.setFont("Helvetica-Bold", 9)
        x = 50
        for i, h in enumerate(headers):
            c.drawString(x, y, h)
            x += col_widths[i]
        y -= 15
        c.line(50, y + 5, width - 50, y + 5)
        c.setFont("Helvetica", 9)
        for idx, p in enumerate(top):
            if y < 50:
                c.showPage()
                y = height - 50
            x = 50
            vals = [str(idx + 1), p.get("name", ""), str(p.get("quantity", 0)),
                    f"₡{p.get('total', 0):,.2f}"]
            for i, v in enumerate(vals):
                c.drawString(x, y, str(v)[:50])
                x += col_widths[i]
            y -= 13

    cats = data.get("categories") or []
    if cats:
        y -= 20
        if y < 100:
            c.showPage()
            y = height - 50
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Ventas por Categoría")
        y -= 20
        c.setFont("Helvetica", 9)
        for cat in cats:
            if y < 50:
                c.showPage()
                y = height - 50
            c.drawString(50, y, f"{cat.get('category', '')}: ₡{cat.get('total', 0):,.2f}")
            y -= 13

    c.showPage()
    c.save()
    return filename


# ============================================================
# ANALÍTICA DE COMPRAS — Exportaciones
# ============================================================

def export_purchases_analytics_excel(data, filename=None):
    """Exporta analítica de compras a Excel (múltiples hojas)."""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"analitica_compras_{now_cr().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb = Workbook()
    first_sheet_used = False

    suppliers = data.get("suppliers") or []
    if suppliers:
        ws = wb.active
        ws.title = "Gasto por proveedor"
        _write_sheet(ws, suppliers)
        first_sheet_used = True

    evolution = data.get("evolution") or []
    if evolution:
        if not first_sheet_used:
            ws = wb.active
            ws.title = "Evolución mensual"
            first_sheet_used = True
        else:
            ws = wb.create_sheet("Evolución mensual")
        _write_sheet(ws, evolution)

    days_data = data.get("payment_days") or {}
    by_supplier = days_data.get("by_supplier") or []
    if by_supplier:
        _write_sheet(wb.create_sheet("Días de pago"), by_supplier)

    products = data.get("top_products") or []
    if products:
        _write_sheet(wb.create_sheet("Top productos"), products)

    # Si no hubo datos, al menos guardar con la hoja vacía
    wb.save(filename)
    return filename


def export_purchases_analytics_pdf(data, filename=None, business_name="Mi Negocio"):
    """Exporta analítica de compras a PDF."""
    # FASE 2 — Fix 2.3: ruta absoluta en DATA_DIR/exports
    filename = _resolve_export_filename(
        filename, f"analitica_compras_{now_cr().strftime('%Y%m%d_%H%M%S')}.pdf"
    )

    c = canvas.Canvas(filename, pagesize=landscape(letter))
    width, height = landscape(letter)
    y = height - 50

    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, y, f"Analítica de Compras - {business_name}")
    y -= 25
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Periodo: {data.get('start_date', '')} al {data.get('end_date', '')}")
    y -= 15
    c.drawString(50, y, f"Generado: {now_cr().strftime('%d/%m/%Y %H:%M')}")
    y -= 30

    suppliers = data.get("suppliers") or []
    if suppliers:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Gasto por Proveedor")
        y -= 20
        headers = ["Proveedor", "Facturas", "Gasto total (₡)", "Promedio (₡)"]
        col_w = [200, 80, 120, 120]
        c.setFont("Helvetica-Bold", 9)
        x = 50
        for i, h in enumerate(headers):
            c.drawString(x, y, h)
            x += col_w[i]
        y -= 15
        c.line(50, y + 5, width - 50, y + 5)
        c.setFont("Helvetica", 9)
        for s in suppliers:
            if y < 50:
                c.showPage()
                y = height - 50
            x = 50
            vals = [s.get("supplier_name", "")[:30], str(s.get("invoice_count", 0)),
                    f"₡{s.get('total_spent', 0):,.2f}", f"₡{s.get('avg_invoice', 0):,.2f}"]
            for i, v in enumerate(vals):
                c.drawString(x, y, v)
                x += col_w[i]
            y -= 13

    products = data.get("top_products") or []
    if products:
        y -= 20
        if y < 100:
            c.showPage()
            y = height - 50
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Top Productos Comprados")
        y -= 20
        headers = ["Producto", "Cantidad", "Gasto (₡)", "Proveedor"]
        col_w = [200, 80, 120, 200]
        c.setFont("Helvetica-Bold", 9)
        x = 50
        for i, h in enumerate(headers):
            c.drawString(x, y, h)
            x += col_w[i]
        y -= 15
        c.line(50, y + 5, width - 50, y + 5)
        c.setFont("Helvetica", 9)
        for p in products:
            if y < 50:
                c.showPage()
                y = height - 50
            x = 50
            vals = [p.get("product_name", "")[:30], str(p.get("total_qty", 0)),
                    f"₡{p.get('total_spent', 0):,.2f}", (p.get("top_supplier") or "—")[:30]]
            for i, v in enumerate(vals):
                c.drawString(x, y, v)
                x += col_w[i]
            y -= 13

    c.showPage()
    c.save()
    return filename