# app/utils/cabys_updater.py
"""
FASE 3 — Fix 3.2: Lectura de Excel CABYS con openpyxl puro.

Se eliminó la dependencia de pandas para leer el catálogo CABYS
del Banco Central.  openpyxl load_workbook(read_only=True) es
más liviano y suficiente para este caso de uso.
"""
import requests
import io
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.core.logger import logger
from app.utils.dt import now_cr

URL_CABYS = "https://www.bccr.fi.cr/indicadores-economicos/cabys/Catalogo-de-bienes-servicios.xlsx"


def update_cabys(db: Session) -> int:
    """
    Actualiza el catálogo CABYS usando la sesión del POS (db).
    Devuelve la cantidad de registros cargados.
    """

    logger.info(f"🔄 Descargando CABYS desde Hacienda... ({now_cr().strftime('%Y-%m-%d %H:%M')})")

    # Descargar archivo Excel
    response = requests.get(URL_CABYS, timeout=20)
    response.raise_for_status()

    # Leer con openpyxl (sin pandas)
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        ws = wb.worksheets[0]  # Hoja 0 (Catálogo)
        logger.info("Hoja 0 ('Catálogo') cargada correctamente.")
    except Exception as e:
        logger.error(f"Error al cargar la hoja 0 del Excel: {e}")
        raise Exception(f"No se pudo leer la hoja 'Catálogo' del Excel: {e}")

    # Leer filas como lista (row 1 = fila 1 de Excel, row 2 = headers)
    rows_iter = ws.iter_rows(min_row=1, values_only=True)

    # Fila 1: título o encabezado general — saltar
    try:
        next(rows_iter)
    except StopIteration:
        raise Exception("El archivo CABYS está vacío.")

    # Fila 2: headers reales
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        raise Exception("El archivo CABYS no tiene fila de encabezados.")

    headers = [str(h).strip().lower() if h else "" for h in raw_headers]

    # Identificar índices de columnas
    code_idx = None
    desc_idx = None
    iva_idx = None
    for i, h in enumerate(headers):
        if h == "categoría 9":
            code_idx = i
        elif h == "descripción (categoría 9)":
            desc_idx = i
        elif h == "impuesto":
            iva_idx = i

    if code_idx is None or desc_idx is None:
        logger.error("❌ No se encontraron columnas válidas en el archivo CABYS.")
        logger.error(
            f"Encontradas: Código idx={code_idx}, Descripción idx={desc_idx}, "
            f"IVA idx={iva_idx}. Disponibles: {headers}"
        )
        raise Exception(
            "No se encontraron columnas válidas. Revise 'Categoría 9', "
            "'Descripción (categoría 9)' e 'Impuesto'."
        )

    # Ajustes
    MAX_DESC_LENGTH = 500  # límite para evitar truncamiento en MySQL
    data = []
    row_count = 0

    for row in rows_iter:
        row_count += 1
        # Código completo (con puntos)
        code = str(row[code_idx] or "").strip() if code_idx < len(row) else ""

        # Descripción truncada
        desc_raw = str(row[desc_idx] or "").strip() if desc_idx < len(row) else ""
        desc = desc_raw[:MAX_DESC_LENGTH]

        # IVA
        iva_raw = str(row[iva_idx] or "0.13").replace("%", "").strip() if (iva_idx is not None and iva_idx < len(row)) else "0.13"

        try:
            iva_float = float(iva_raw)
            # Si viene 0.13 -> convertir a 13
            iva_calculated = iva_float * 100 if iva_float < 1 else iva_float
            iva = int(iva_calculated)
        except ValueError:
            iva = 13  # valor por defecto

        # Tomar solo códigos finales (13 dígitos quitando puntos)
        if code and desc and len(code.replace(".", "")) == 13:
            data.append((code, desc, iva))

    # Cerrar workbook read_only
    wb.close()

    logger.info(f"📄 {row_count} filas leídas del archivo CABYS, {len(data)} registros válidos.")

    # Guardar en DB
    logger.info("🧹 Limpiando tabla CABYS...")
    db.execute(text("DELETE FROM cabys;"))

    logger.info("💾 Insertando registros CABYS...")
    insert_data = [{"code": c, "description": d, "iva": i} for (c, d, i) in data]

    db.execute(
        text("INSERT INTO cabys (code, description, iva) VALUES (:code, :description, :iva)"),
        insert_data
    )
    db.commit()

    logger.info(f"✅ CABYS actualizado correctamente ({len(data)} registros).")

    return len(data)